"""Excel export functionality for speaker content"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import config


class ExcelExporter:
    """Export processed speaker data to Excel format"""

    def __init__(self):
        self.wb = Workbook()

    def export(self, speakers_data, output_path):
        """
        Export speaker data to Excel file.

        Args:
            speakers_data: List of processed speaker data dicts
            output_path: Path to save the Excel file

        Returns:
            str: Path to the created file
        """
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create main content sheet
        self._create_content_sheet(speakers_data)

        # Create quality control sheet
        self._create_qc_sheet(speakers_data)

        # Save the workbook
        self.wb.save(output_path)
        return output_path

    def _create_content_sheet(self, speakers_data):
        """Create the main content sheet with formatted speaker data."""
        ws = self.wb.create_sheet("Speaker Content")

        # Define headers
        headers = [
            "Speaker Name",
            "Bio (50 words)",
            "Bio (100 words)",
            "Emcee Intro",
            "Session Title",
            "Session Abstract (75 words)",
            "Key Takeaway 1",
            "Key Takeaway 2",
            "Key Takeaway 3",
            "Headshot Alt Text",
            "Tech Requirements"
        ]

        # Style the header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add speaker data
        for row_num, speaker in enumerate(speakers_data, 2):
            processed = speaker.get('processed', {})
            bio = processed.get('bio', {})
            session = processed.get('session', {})
            takeaways = session.get('takeaways', ['', '', ''])

            # Ensure we have 3 takeaways
            while len(takeaways) < 3:
                takeaways.append('')

            row_data = [
                speaker.get('name', ''),
                bio.get('short', ''),
                bio.get('medium', ''),
                bio.get('intro', ''),
                speaker.get('session_title', ''),
                session.get('abstract', ''),
                takeaways[0],
                takeaways[1],
                takeaways[2],
                processed.get('alt_text', ''),
                speaker.get('tech_requirements', '')
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Set column widths
        column_widths = {
            'A': 20,  # Speaker Name
            'B': 35,  # Bio 50
            'C': 45,  # Bio 100
            'D': 40,  # Emcee Intro
            'E': 30,  # Session Title
            'F': 45,  # Abstract
            'G': 35,  # Takeaway 1
            'H': 35,  # Takeaway 2
            'I': 35,  # Takeaway 3
            'J': 40,  # Alt Text
            'K': 30   # Tech Reqs
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze the header row
        ws.freeze_panes = 'A2'

    def _create_qc_sheet(self, speakers_data):
        """Create the quality control sheet."""
        ws = self.wb.create_sheet("Quality Control")

        # Style the header row
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        headers = [
            "Speaker Name",
            "Headshot Present",
            "Headshot Valid",
            "Tech Requirements",
            "Missing Tech Items",
            "Session Description Clear",
            "Vague Language",
            "Bio Word Count",
            "Bio Under 500 Words",
            "Name Mismatch",
            "Buzzwords Found",
            "Issues",
            "Warnings"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add QC data
        for row_num, speaker in enumerate(speakers_data, 2):
            qc = speaker.get('quality_control', {})
            checklist = qc.get('checklist', {})

            # Format boolean values as Yes/No
            def bool_to_yn(val):
                return "Yes" if val else "No"

            row_data = [
                speaker.get('name', ''),
                bool_to_yn(checklist.get('headshot_present', False)),
                bool_to_yn(checklist.get('headshot_valid', False)),
                bool_to_yn(checklist.get('tech_requirements_specified', False)),
                ', '.join(checklist.get('missing_tech_items', [])),
                bool_to_yn(checklist.get('session_description_clear', True)),
                ', '.join(checklist.get('vague_language_detected', [])),
                checklist.get('bio_word_count', 0),
                bool_to_yn(checklist.get('bio_under_limit', True)),
                bool_to_yn(checklist.get('name_mismatch', False)),
                ', '.join(checklist.get('buzzwords_found', [])),
                ' | '.join(qc.get('issues', [])),
                ' | '.join(qc.get('warnings', []))
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Color code issues and warnings
                if col_num == len(headers) - 1 and value:  # Issues column
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif col_num == len(headers) and value:  # Warnings column
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # Set column widths
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        # Make issues and warnings columns wider
        ws.column_dimensions[get_column_letter(len(headers) - 1)].width = 40
        ws.column_dimensions[get_column_letter(len(headers))].width = 40

        # Freeze the header row
        ws.freeze_panes = 'A2'

    def add_metadata_sheet(self, metadata):
        """Add a metadata sheet with processing information."""
        ws = self.wb.create_sheet("Processing Info", 0)

        ws['A1'] = "Mini AI Speaker Agent - Processing Report"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws['A4'] = "Total Speakers Processed:"
        ws['B4'] = metadata.get('total_speakers', 0)

        ws['A5'] = "Speakers with Issues:"
        ws['B5'] = metadata.get('speakers_with_issues', 0)

        ws['A6'] = "Speakers with Warnings:"
        ws['B6'] = metadata.get('speakers_with_warnings', 0)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
